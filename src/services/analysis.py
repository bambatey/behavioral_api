import zipfile
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.datalayer.repository import ParticipantRepository, TrialResultRepository

from ._design_coding import (
    CONDITION_CODE,
    POSITION_LABELS,
    POSITION_TO_ORDER,
    POSITION_TO_PRONOUN,
    design_factors,
)

if TYPE_CHECKING:
    from firebase_admin.firestore_async import AsyncClient


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# R script that ships next to the Excel in /admin/analysis bundle.
_R_TEMPLATE_PATH = Path(__file__).with_name("analysis_template.R")


class AnalysisService:
    """Per-participant × per-condition aggregations.
    The first sheet ("Long format") is what an R / dplyr user wants — one row
    per trial response with the design factors already decoded."""

    def __init__(self, db: "AsyncClient"):
        self.db = db
        self.participant_repo = ParticipantRepository(db)
        self.trial_repo = TrialResultRepository(db)

    async def export(self) -> bytes:
        participants = {p.id: p for p in await self.participant_repo.get_all()}
        trials = await self.trial_repo.get_all()

        # Critical trials with a usable Likert response — the analysis universe
        critical = [
            t for t in trials
            if not t.is_filler
            and t.position is not None
            and t.bias is not None
            and isinstance(t.response, int)
        ]

        wb = Workbook()

        # ----- Sheet 1: LONG FORMAT (R-friendly) -----
        ws_long = wb.active
        ws_long.title = "Long format"
        self._header(ws_long, [
            "participant", "assignment_index", "trial_index",
            "item_type", "item_id",
            "pronoun", "order", "antecedent", "condition", "congruency",
            "position", "position_label", "bias",
            "rating", "rt", "correct_answer",
        ])
        long_rows = []
        for t in trials:  # include fillers too; R user can filter on item_type
            if not isinstance(t.response, int):
                continue
            p = participants.get(t.participant_id)
            item_type = "filler" if t.is_filler else "critical"
            item_id = t.filler_id if t.is_filler else t.sentence_id
            factors = design_factors(t.position, t.bias)
            long_rows.append([
                p.name if p else t.participant_id,
                p.assignment_index if p else None,
                t.trial_index,
                item_type,
                item_id,
                factors["pronoun"],
                factors["order"],
                factors["antecedent"],
                factors["condition"],
                factors["congruency"],
                t.position,
                factors["position_label"],
                t.bias,
                t.response,
                t.rt,
                t.correct_answer,
            ])
        long_rows.sort(key=lambda r: ((r[0] or "").lower(), r[2] or 0))
        self._write_rows(ws_long, long_rows, start_row=2)

        # ----- Sheet 2: Per (participant × position_label) — 6 rows / participant -----
        ws_pos = wb.create_sheet("By position-label")
        self._header(ws_pos, [
            "Participant ID", "Name", "Assignment Index",
            "Position", "Position Label", "N trials", "Mean Likert",
        ])
        per_pos: dict = defaultdict(list)
        for t in critical:
            per_pos[(t.participant_id, t.position)].append(t.response)
        pos_rows = []
        for (pid, pos), vals in per_pos.items():
            p = participants.get(pid)
            pos_rows.append([
                pid,
                p.name if p else None,
                p.assignment_index if p else None,
                pos,
                POSITION_LABELS.get(pos),
                len(vals),
                round(sum(vals) / len(vals), 2),
            ])
        pos_rows.sort(key=lambda r: ((r[1] or "").lower(), r[3]))
        self._write_rows(ws_pos, pos_rows, start_row=2)

        # ----- Sheet 3: Per (participant × condition) — the 12-cell design -----
        ws_cond = wb.create_sheet("By condition (12-cell)")
        self._header(ws_cond, [
            "Participant ID", "Name", "Assignment Index",
            "Condition", "Pronoun", "Order", "Antecedent",
            "N trials", "Mean Likert",
        ])
        per_cond: dict = defaultdict(list)
        for t in critical:
            factors = design_factors(t.position, t.bias)
            key = (t.participant_id, factors["condition"], factors["pronoun"],
                   factors["order"], factors["antecedent"])
            per_cond[key].append(t.response)
        cond_rows = []
        for (pid, cond, pron, order, ant), vals in per_cond.items():
            p = participants.get(pid)
            cond_rows.append([
                pid,
                p.name if p else None,
                p.assignment_index if p else None,
                cond, pron, order, ant,
                len(vals),
                round(sum(vals) / len(vals), 2),
            ])
        cond_rows.sort(key=lambda r: ((r[1] or "").lower(), r[3] or ""))
        self._write_rows(ws_cond, cond_rows, start_row=2)

        # ----- Sheet 4: Wide format — one row per participant, 12 condition cols -----
        ws_wide = wb.create_sheet("Wide (12 conditions)")
        cond_cols = [code for code, _ in
                     sorted(((code, sort_key(code)) for code in CONDITION_CODE.values()),
                            key=lambda x: x[1])]
        self._header(
            ws_wide,
            ["Participant ID", "Name", "Assignment Index"] + cond_cols,
        )
        # invert per_cond by participant
        wide_table: dict = defaultdict(dict)  # pid -> condCode -> mean
        for (pid, cond, *_rest), vals in per_cond.items():
            wide_table[pid][cond] = round(sum(vals) / len(vals), 2)
        wide_rows = []
        for pid, p in participants.items():
            row = [pid, p.name, p.assignment_index]
            for cond in cond_cols:
                row.append(wide_table.get(pid, {}).get(cond))
            wide_rows.append(row)
        wide_rows.sort(key=lambda r: (r[1] or "").lower())
        self._write_rows(ws_wide, wide_rows, start_row=2)

        # ----- Sheet 5: Group means -----
        ws_grp = wb.create_sheet("Group means")
        self._header(ws_grp, [
            "Condition", "Pronoun", "Order", "Antecedent", "Congruency",
            "N participants", "N trials", "Mean Likert", "SD",
        ])
        group: dict = defaultdict(list)  # (cond, pron, order, ant) -> [ratings]
        group_participants: dict = defaultdict(set)
        for t in critical:
            f = design_factors(t.position, t.bias)
            key = (f["condition"], f["pronoun"], f["order"], f["antecedent"], f["congruency"])
            group[key].append(t.response)
            group_participants[key].add(t.participant_id)
        grp_rows = []
        for key, vals in group.items():
            cond, pron, order, ant, congruency = key
            mean = sum(vals) / len(vals)
            sd = (sum((v - mean) ** 2 for v in vals) / len(vals)) ** 0.5 if len(vals) > 1 else None
            grp_rows.append([
                cond, pron, order, ant, congruency,
                len(group_participants[key]),
                len(vals),
                round(mean, 2),
                round(sd, 2) if sd is not None else None,
            ])
        grp_rows.sort(key=lambda r: sort_key(r[0] or "Z99"))
        self._write_rows(ws_grp, grp_rows, start_row=2)

        # Auto-width every sheet
        for ws in [ws_long, ws_pos, ws_cond, ws_wide, ws_grp]:
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
            ws.freeze_panes = "A2"

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    async def export_bundle(self) -> bytes:
        """Zip of the analysis Excel + a ready-to-run R script.

        Layout:
          analysis.xlsx  -- five-sheet workbook (Long format first)
          analysis.R     -- descriptives + clmm/lmer + emmeans + ggplot pipeline
        """
        excel_bytes = await self.export()
        r_script = _R_TEMPLATE_PATH.read_text(encoding="utf-8")

        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("analysis.xlsx", excel_bytes)
            zf.writestr("analysis.R", r_script)
        buf.seek(0)
        return buf.getvalue()

    @staticmethod
    def _header(ws, titles):
        for col_num, title in enumerate(titles, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = title
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _write_rows(ws, rows, start_row: int = 2):
        for r_idx, row in enumerate(rows, start=start_row):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx).value = val


def sort_key(code: str) -> int:
    """C1, C2, ..., C12 → numeric for ordering."""
    if not code or not code.startswith("C"):
        return 9999
    try:
        return int(code[1:])
    except ValueError:
        return 9999
