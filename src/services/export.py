from io import BytesIO
from typing import List, Optional, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.datalayer.model import TrialResult
from src.datalayer.repository import ParticipantRepository, TrialResultRepository

from ._design_coding import design_factors

if TYPE_CHECKING:
    from firebase_admin.firestore_async import AsyncClient


# Long-format columns — one row per response. Decoded design factors (pronoun,
# order, antecedent, condition, congruency, position_label) sit next to the raw
# (bias, position) so R / SPSS / JASP can pivot however they want.
COLUMNS = [
    "Participant ID",
    "Participant Name",
    "Assignment Index",
    "Trial Index",
    "Item Type",       # 'critical' | 'filler'
    "Item ID",         # sentence_id for critical, filler_id for filler
    "Context ID",
    "Context Text",
    "Sentence Text",
    # raw stimulus fields
    "Bias",
    "Position",
    # decoded design factors
    "Pronoun",
    "Order",
    "Antecedent",
    "Condition",
    "Congruency",
    "Position Label",
    # response
    "Response (Likert 1-7)",
    "Correct Answer",
    "RT (ms)",
    "Created At",
]


class ExportService:
    """Long-format Excel export of all (or one participant's) trial results.
    One row per response, with design factors pre-decoded."""

    def __init__(self, db: "AsyncClient"):
        self.db = db
        self.participant_repo = ParticipantRepository(db)
        self.trial_repo = TrialResultRepository(db)

    async def export_to_excel(self, participant_id: Optional[str] = None) -> bytes:
        if participant_id:
            trials = await self.trial_repo.find_by_participant_id(participant_id)
        else:
            trials = await self.trial_repo.get_all()
        participants = {p.id: p for p in await self.participant_repo.get_all()}
        return self._generate_excel(trials, participants)

    def _generate_excel(self, trials: List[TrialResult], participants_by_id: dict) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trials (long)"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, title in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sort by participant, then by trial_index — easier on humans browsing the file.
        trials_sorted = sorted(
            trials,
            key=lambda t: (t.participant_id or "", t.trial_index if t.trial_index is not None else 0),
        )

        for row_num, trial in enumerate(trials_sorted, 2):
            participant = participants_by_id.get(trial.participant_id)
            assignment_index = participant.assignment_index if participant else None
            item_type = "filler" if trial.is_filler else "critical"
            item_id = trial.filler_id if trial.is_filler else trial.sentence_id
            factors = design_factors(trial.position, trial.bias)

            row_data = [
                trial.participant_id,
                trial.participant_name,
                assignment_index,
                trial.trial_index,
                item_type,
                item_id,
                trial.context_id,
                trial.context_text,
                trial.sentence_text,
                trial.bias,
                trial.position,
                factors["pronoun"],
                factors["order"],
                factors["antecedent"],
                factors["condition"],
                factors["congruency"],
                factors["position_label"],
                trial.response,
                trial.correct_answer,
                trial.rt,
                trial.created_at.isoformat() if trial.created_at else None,
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
